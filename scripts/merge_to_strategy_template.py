#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Safe-Append Merge for Strategy Template (keeps formulas & charts intact)

特性：
- 不修改首行表头，不删除任何已存在的行与单元格（避免破坏公式/图表/表格结构）
- 仅将 CSV 的“数据行”追加到工作表末尾
- OKLO/LEU：按首列键（通常是日期）去重；Dashboard：不去重
- 自动将 pandas 的 NaN/NaT 写成 None，避免 Excel 打开报错
- 若目标 Sheet 完全空白（无表头），首次写入会把 CSV 的表头写到第 1 行

映射关系：
  OKLO      <- OKLO_Daily.csv         （去重：首列）
  LEU       <- LEU_Daily.csv          （去重：首列）
  Dashboard <- All_Signals.csv        （不去重）

用法：
  python merge_to_strategy_template.py --template 10-25-OKLO_LEU_Strategy_Template.xlsx
"""

import argparse
import os
import pandas as pd
from openpyxl import load_workbook

# ---------------- Helpers ----------------

def safe_read_csv(path: str) -> pd.DataFrame:
    """安全读取 CSV（UTF-8），读取失败返回空 DF。"""
    if not os.path.exists(path):
        print(f"⚠️ 找不到 {path}，跳过")
        return pd.DataFrame()
    try:
        df = pd.read_csv(path, encoding="utf-8")
        print(f"✅ 读取 {path}（{len(df)} 行，{len(df.columns)} 列）")
        return df
    except Exception as e:
        print(f"❌ 读取 {path} 失败：{e}")
        return pd.DataFrame()

def get_headers(ws) -> list:
    """读取首行表头（不修改）。若为空表，返回空列表。"""
    if (ws.max_row or 0) < 1:
        return []
    headers = []
    max_col = ws.max_column or 0
    for c in range(1, max_col + 1):
        headers.append(ws.cell(row=1, column=c).value)
    # 去掉末尾的空表头
    while headers and (headers[-1] is None or str(headers[-1]).strip() == ""):
        headers.pop()
    return [str(h) if h is not None else "" for h in headers]

def align_df_to_headers(df: pd.DataFrame, headers: list) -> pd.DataFrame:
    """
    严格按现有 headers 对齐；缺列补 NA；多余列丢弃（保护模板结构，避免新增列导致图表/公式被破坏）。
    若 headers 为空（即空白 Sheet），直接返回 df（外层会先写表头）。
    """
    if not headers:
        return df.copy()
    df2 = df.copy()
    df2.columns = [str(c) for c in df2.columns]
    # 缺的列补 NA
    for col in headers:
        if col not in df2.columns:
            df2[col] = pd.NA
    # 多余列丢弃
    df2 = df2[headers]
    return df2

def normalize_cell_value(v):
    """将 pandas 的 NA/NaN/NaT 统一为 None，其余原样返回。"""
    if pd.isna(v):
        return None
    return v

def safe_write_cell(ws, r, c, value):
    """保护性写入，确保 row/col 为 >=1 的整数，避免 NoneType 导致 openpyxl 报错。"""
    try:
        r = int(r) if r is not None and int(r) >= 1 else 1
    except Exception:
        r = 1
    try:
        c = int(c) if c is not None and int(c) >= 1 else 1
    except Exception:
        c = 1
    ws.cell(row=r, column=c, value=value)

def collect_existing_keys(ws, key_col=1) -> set:
    """
    收集用于去重的“已存在键集”（默认取第 1 列）。
    若 key_col 为 None，则返回空集合（表示不启用去重）。
    """
    if key_col is None:
        return set()
    keys = set()
    max_row = ws.max_row or 1
    for r in range(2, max_row + 1):
        v = ws.cell(row=r, column=key_col).value
        if v is None:
            continue
        keys.add(str(v))
    return keys

def append_df(ws, df: pd.DataFrame, key_col=1) -> int:
    """
    仅追加 df 的数据行；不修改首行表头，不删除任何行。
    - 若 key_col 为 None：不去重
    - 若 key_col 为整数：按对应列（1 起）做去重
    返回：本次追加的行数
    """
    if df is None or df.empty:
        return 0

    # 将 df 中的 NA/NaN 转为 None，并转二维列表
    df_to_write = df.where(pd.notna(df), None)
    rows = [list(row) for row in df_to_write.itertuples(index=False, name=None)]

    existing = collect_existing_keys(ws, key_col=key_col)
    added = 0

    for row_vals in rows:
        # 去重判断（仅当指定 key_col 时）
        if key_col is not None and key_col >= 1:
            if len(row_vals) >= key_col:
                key = row_vals[key_col - 1]
                key_str = "" if key is None else str(key)
                if key_str and key_str in existing:
                    # 已存在该键（如日期），跳过
                    continue

        # 计算追加的位置：始终写到末尾下一行
        max_row = ws.max_row or 1
        start_row = max_row + 1 if max_row >= 1 else 1

        # 逐列安全写入
        for c_idx, value in enumerate(row_vals, start=1):
            safe_write_cell(ws, start_row, c_idx, normalize_cell_value(value))

        # 更新已存在键集
        if key_col is not None and key_col >= 1 and len(row_vals) >= key_col:
            key = row_vals[key_col - 1]
            if key is not None:
                existing.add(str(key))

        added += 1

    return added

# ---------------- Merge Core ----------------

def merge_sheet(wb, sheet_name: str, csv_path: str, key_is_first_col: bool = True) -> int:
    """
    将 csv_path 对应的 CSV 追加写入到工作簿的 sheet_name。
    - 若 Sheet 无表头：首次写入会把 CSV 的表头写到第 1 行
    - key_is_first_col=True  -> 用首列去重（OKLO/LEU）
      key_is_first_col=False -> 不去重（Dashboard）
    返回：追加的行数
    """
    df = safe_read_csv(csv_path)
    if df.empty:
        return 0

    if sheet_name not in wb.sheetnames:
        print(f"⚠️ 模板不存在 Sheet: {sheet_name}，跳过")
        return 0

    ws = wb[sheet_name]
    headers = get_headers(ws)

    if not headers:
        # 空白 Sheet：写入 CSV 表头（只在第一次发生）
        for idx, name in enumerate(list(df.columns), start=1):
            safe_write_cell(ws, 1, idx, str(name))
        headers = [str(c) for c in df.columns]
        print(f"🧭 Sheet `{sheet_name}` 初次写入表头：{headers}")

    # 严格对齐到现有表头（不新增列、不修改表头）
    df_aligned = align_df_to_headers(df, headers)

    # 仅追加，不清空、不改表头
    key_col = 1 if key_is_first_col else None
    added = append_df(ws, df_aligned, key_col=key_col)
    print(f"✔ Sheet `{sheet_name}` 追加 {added} 行")
    return added

def merge_to_template(template_path: str):
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"未找到模板：{template_path}")

    wb = load_workbook(template_path)

    total_added = 0
    total_added += merge_sheet(wb, "OKLO", "OKLO_Daily.csv", key_is_first_col=True)
    total_added += merge_sheet(wb, "LEU",  "LEU_Daily.csv", key_is_first_col=True)
    total_added += merge_sheet(wb, "Dashboard", "All_Signals.csv", key_is_first_col=False)

    out_path = os.path.splitext(template_path)[0] + "_updated.xlsx"
    wb.save(out_path)
    print(f"🎯 已保存：{out_path}（本次共追加 {total_added} 行；未触碰表头/公式）")

# ---------------- CLI ----------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", required=True,
                        help="模板文件路径（例如 10-25-OKLO_LEU_Strategy_Template.xlsx）")
    args = parser.parse_args()
    merge_to_template(args.template)

if __name__ == "__main__":
    main()
